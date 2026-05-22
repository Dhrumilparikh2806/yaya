import csv
import io

import openpyxl

from app.processors.base import BaseProcessor

_ROW_LIMIT = 500


class XLSXProcessor(BaseProcessor):
    def extract(self) -> str:
        if self.job.file_path.endswith(".csv"):
            return self._extract_csv()
        return self._extract_xlsx()

    def _extract_csv(self) -> str:
        with open(self.job.file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)

        if len(rows) > _ROW_LIMIT + 1:
            self.log.warning("csv_rows_truncated", total=len(rows), limit=_ROW_LIMIT)
            rows = rows[:_ROW_LIMIT + 1]

        return self._table_to_markdown(rows)

    def _extract_xlsx(self) -> str:
        wb = openpyxl.load_workbook(self.job.file_path, read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [[str(cell.value) if cell.value is not None else "" for cell in row] for row in ws.iter_rows()]
            if len(rows) > _ROW_LIMIT + 1:
                self.log.warning("xlsx_rows_truncated", sheet=sheet_name, total=len(rows), limit=_ROW_LIMIT)
                rows = rows[:_ROW_LIMIT + 1]
            parts.append(f"[Sheet: {sheet_name}]\n" + self._table_to_markdown(rows))
        return "\n\n".join(parts)

    def summarise(self, text: str, db) -> dict:
        if len(text) > 30000:
            self.log.warning("xlsx_text_truncated", original_len=len(text), truncated_to=30000)
            text = text[:30000]

        prompt = f"""You are a data analyst. Analyse the following spreadsheet data and return ONLY valid JSON.
No preamble, no markdown code blocks, just raw JSON.

Return this exact structure:
{{
  "title": "spreadsheet title or filename",
  "summary": "2-3 sentence summary of what this data contains",
  "sheets": ["list of sheet names found"],
  "column_descriptions": {{"column_name": "what it likely represents"}},
  "key_insights": ["notable patterns, max/min values, trends noticed"],
  "row_count": 0
}}

Spreadsheet data:
{text}
"""
        return self._call_gemini_json(prompt, db)
